from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from database import SessionLocal, engine
from sqlalchemy.orm import Session
from models import DocumentType, UploadedFile, ExtractData


session = SessionLocal()
rows = session.query(ExtractData).order_by(
    ExtractData.title,
    ExtractData.timestamp
).all()

grouped = defaultdict(list)

for row in rows:
    grouped[row.title].append(row)

wb = Workbook()

wb.remove(wb.active)

for title, items in grouped.items():

    sheet_name = title[:31]
    ws = wb.create_sheet(title=sheet_name)

    ws["A1"] = "Дата"
    ws["B1"] = "Value"

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    for row_num, item in enumerate(items, start=2):

        ws.cell(row=row_num, column=1, value=item.timestamp)

        try:
            value = float(item.value.replace(",", "."))
        except Exception:
            value = None

        ws.cell(row=row_num, column=2, value=value)

    for cell in ws["A"][1:]:
        cell.number_format = "dd.mm.yyyy hh:mm"

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Дата"

    data = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=len(items) + 1
    )

    dates = Reference(
        ws,
        min_col=1,
        min_row=2,
        max_row=len(items) + 1
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)

    chart.height = 10
    chart.width = 20

    ws.add_chart(chart, "D2")

wb.save("extractdata_report.xlsx")