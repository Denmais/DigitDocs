import os

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


# Пример выгрузки по title
def excel(db: Session, args):
    task_id = args.get("task_id")

    data = (
        db.query(ExtractData)
        .filter(ExtractData.upload_id == task_id)
        .all()
    )

    if not data:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Выгрузка"

    ws.append(["Параметр", "Итоговое значение"])

    for item in data:
        ws.append([item.ru_title, item.value])

    pdf_name = os.path.basename(data[0].doc_name)

    chart_data = (
        db.query(ExtractData)
        .filter(ExtractData.doc_name.like(f"%/{pdf_name}"))
        .filter(ExtractData.title == "tariff_kw_night")
        .order_by(ExtractData.timestamp)
        .all()
    )

    if chart_data:
        chart_ws = wb.create_sheet("График")

        chart_ws.append(["timestamp", "value"])

        for row in chart_data:
            try:
                value = float(row.value.replace(",", "."))
            except Exception:
                continue

            chart_ws.append([
                row.timestamp,
                value,
            ])

        chart = LineChart()
        chart.title = "tariff_kw_night"
        chart.y_axis.title = "value"
        chart.x_axis.title = "timestamp"

        data_ref = Reference(
            chart_ws,
            min_col=2,
            min_row=1,
            max_row=chart_ws.max_row
        )

        cats_ref = Reference(
            chart_ws,
            min_col=1,
            min_row=2,
            max_row=chart_ws.max_row
        )

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart.height = 10
        chart.width = 20

        chart_ws.add_chart(chart, "E2")

    path = f"{XLS}/{task_id}.xlsx"
    wb.save(path)

    return task_id